"""Deterministic reporting formulas and file-based report generation."""

from __future__ import annotations

import csv
import json
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook

from backend.netgenix.reports.pdf import write_pdf_report


PROJECT_ROOT = Path(__file__).resolve().parents[3]
REPORT_ROOT = PROJECT_ROOT / "runtime" / "reports"

CONCEPT_ALIASES: dict[str, list[str]] = {
    "site_name": ["site_name", "site", "site id", "eNodeB Name", "Node Name"],
    "traffic_gb": ["traffic_gb", "traffic gb", "data traffic gb", "4g traffic gb", "total traffic gb", "Total Traffic (Gbit)"],
    "prb_busy_hour": ["prb_busy_hour", "busy hour prb", "prb utilization", "dl prb utilization", "DL PRB Usage Rate(%)"],
    "code_drop": ["code_drop", "code drop", "call drop", "drop rate", "erab drop rate", "Call Drop Rate (All)(%)", "Call Drop Rate (All)"],
    "radio_network_availability": [
        "radio_network_availability",
        "radio network availability",
        "network availability",
        "availability",
        "rna",
        "Radio Net Availability Rate(%)",
    ],
    "active_subscribers": ["active_subscribers", "active users", "active subscriber", "users"],
    "addressable_subscribers": ["addressable_subscribers", "subscriber base", "total subscribers", "potential subscribers"],
    "peak_throughput_mbps": ["peak_throughput_mbps", "peak throughput", "peak throughput mbps", "max throughput mbps"],
    "total_throughput_mbps": [
        "total_throughput_mbps",
        "throughput mbps",
        "average throughput mbps",
        "dl throughput mbps",
        "User DL PDCP Average Throughput",
    ],
}


def _clean(values: Iterable[float | int | None]) -> list[float]:
    return [float(value) for value in values if value is not None]


def average(values: Iterable[float | int | None]) -> float:
    clean = _clean(values)
    return sum(clean) / len(clean) if clean else 0.0


def gb_to_tb(gb: float | int | None) -> float:
    return round(float(gb or 0.0) / 1000.0, 4)


def weekly_traffic_total_gb(daily_gb: Iterable[float | int | None]) -> float:
    return round(sum(_clean(daily_gb)), 4)


def weekly_traffic_total_tb(daily_gb: Iterable[float | int | None]) -> float:
    return gb_to_tb(weekly_traffic_total_gb(daily_gb))


def prb_busy_hour_weekly_average(busy_hour_prb_values: Iterable[float | int | None]) -> float:
    return round(average(busy_hour_prb_values), 4)


def code_drop_average(code_drop_values: Iterable[float | int | None]) -> float:
    return round(average(code_drop_values), 4)


def penetration_rate(active_subscribers: float | int, addressable_subscribers: float | int) -> float:
    denominator = float(addressable_subscribers or 0.0)
    if denominator == 0:
        return 0.0
    return round((float(active_subscribers or 0.0) / denominator) * 100.0, 4)


def average_gb_per_active_user(total_traffic_gb: float | int, active_users: float | int) -> float:
    users = float(active_users or 0.0)
    if users == 0:
        return 0.0
    return round(float(total_traffic_gb or 0.0) / users, 4)


def average_throughput_per_active_user(total_throughput_mbps: float | int, active_users: float | int) -> float:
    users = float(active_users or 0.0)
    if users == 0:
        return 0.0
    return round(float(total_throughput_mbps or 0.0) / users, 4)


def maximum(values: Iterable[float | int | None]) -> float:
    clean = _clean(values)
    return max(clean) if clean else 0.0


def rank_sites(
    records: Sequence[Mapping[str, object]],
    metric: str,
    *,
    top_n: int = 20,
    exclusions: Iterable[str] | None = None,
    descending: bool = True,
) -> list[dict[str, object]]:
    excluded = {site.strip().lower() for site in (exclusions or [])}
    eligible = [
        dict(record)
        for record in records
        if str(record.get("site_name", "")).strip().lower() not in excluded
        and record.get(metric) is not None
    ]

    return sorted(
        eligible,
        key=lambda record: float(record.get(metric) or 0.0),
        reverse=descending,
    )[:top_n]


@dataclass(frozen=True)
class ReportAuditRecord:
    raw_input_files: list[str]
    computed_metrics: dict[str, float]
    generated_output: str | None = None
    user_context: str = "local"
    created_at: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


def _normalise_key(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def _first(row: Mapping[str, object], candidates: Sequence[str], default: object = None) -> object:
    normalised = {_normalise_key(key): value for key, value in row.items()}
    for candidate in candidates:
        key = _normalise_key(candidate)
        if key in normalised and normalised[key] not in ("", None):
            return normalised[key]
    return default


def _matched_column(columns: Sequence[str], candidates: Sequence[str]) -> str | None:
    normalised = {_normalise_key(column): column for column in columns}
    for candidate in candidates:
        key = _normalise_key(candidate)
        if key in normalised:
            return normalised[key]
    return None


def _number(value: object) -> float:
    if value in ("", None):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def _maybe_number(value: object) -> float | None:
    if value in ("", None):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _looks_like_header(values: Sequence[object]) -> bool:
    normalised = {_normalise_key(str(value or "")) for value in values}
    return (
        "date" in normalised
        and (
            "enodebname" in normalised
            or "wholenetwork" in normalised
            or "sitename" in normalised
            or "site" in normalised
        )
    )


def _rows_to_dicts(raw_rows: Sequence[Sequence[object]]) -> list[dict[str, object]]:
    header_index = next(
        (index for index, row in enumerate(raw_rows) if _looks_like_header(row)),
        None,
    )
    if header_index is None:
        return []

    headers = [str(header).strip() if header is not None else "" for header in raw_rows[header_index]]
    output: list[dict[str, object]] = []
    for row in raw_rows[header_index + 1:]:
        if not any(value not in ("", None) for value in row):
            continue
        output.append({
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        })
    return output


def _metric_values(
    rows: Sequence[Mapping[str, object]],
    concept: str,
    *,
    convert_gbit_to_gb: bool = False,
) -> list[float]:
    values: list[float] = []
    aliases = CONCEPT_ALIASES[concept]

    for row in rows:
        matched_column = _matched_column(list(row.keys()), aliases)
        if not matched_column:
            continue

        value = _maybe_number(row.get(matched_column))
        if value is None:
            continue

        if convert_gbit_to_gb and "gbit" in matched_column.lower():
            value = value / 8.0

        values.append(value)

    return values


def _deduplicate_rows(rows: Sequence[Mapping[str, object]]) -> list[dict[str, object]]:
    seen: set[tuple[tuple[str, str], ...]] = set()
    output: list[dict[str, object]] = []

    for row in rows:
        fingerprint = tuple(
            sorted(
                (
                    str(key),
                    str(value).strip(),
                )
                for key, value in row.items()
                if key != "_source_file" and value not in ("", None)
            )
        )
        if fingerprint in seen:
            continue
        seen.add(fingerprint)
        output.append(dict(row))

    return output


def load_tabular_file(path: Path) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return _rows_to_dicts(list(csv.reader(handle)))

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        output: list[dict[str, object]] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            output.extend(_rows_to_dicts(rows))
        return output

    raise ValueError(f"Unsupported report input type: {suffix}")


def preview_column_mapping(path: Path, *, original_filename: str) -> dict[str, object]:
    rows = load_tabular_file(path)
    columns = list(rows[0].keys()) if rows else []
    required = {"site_name", "traffic_gb"}
    mappings = []

    for concept, aliases in CONCEPT_ALIASES.items():
        matched = _matched_column(columns, aliases)
        mappings.append({
            "concept": concept,
            "matched_column": matched,
            "confidence": "high" if matched else "missing",
            "required": concept in required,
        })

    warnings = [
        f"Missing required column mapping: {mapping['concept']}"
        for mapping in mappings
        if mapping["required"] and not mapping["matched_column"]
    ]

    return {
        "filename": original_filename,
        "row_count": len(rows),
        "columns": columns,
        "mappings": mappings,
        "warnings": warnings,
    }


def compute_site_metrics(
    rows: Sequence[Mapping[str, object]],
    *,
    exclusions: Iterable[str] | None = None,
) -> list[dict[str, object]]:
    excluded = {site.strip().lower() for site in (exclusions or []) if site.strip()}
    grouped: dict[str, list[Mapping[str, object]]] = {}

    for row in rows:
        site_name = str(_first(row, CONCEPT_ALIASES["site_name"], "")).strip()
        if not site_name:
            continue
        grouped.setdefault(site_name, []).append(row)

    metrics: list[dict[str, object]] = []
    for site_name, site_rows in grouped.items():
        traffic_values = _metric_values(site_rows, "traffic_gb", convert_gbit_to_gb=True)
        prb_values = _metric_values(site_rows, "prb_busy_hour")
        code_drop_values = _metric_values(site_rows, "code_drop")
        availability_values = _metric_values(site_rows, "radio_network_availability")
        active_subscribers = average(_metric_values(site_rows, "active_subscribers"))
        addressable_subscribers = average(_metric_values(site_rows, "addressable_subscribers"))
        total_throughput = average(_metric_values(site_rows, "total_throughput_mbps"))
        peak_throughput = maximum(_metric_values(site_rows, "peak_throughput_mbps"))
        if peak_throughput == 0:
            peak_throughput = total_throughput
        weekly_gb = weekly_traffic_total_gb(traffic_values)

        metrics.append({
            "site_name": site_name,
            "weekly_traffic_gb": weekly_gb,
            "weekly_traffic_tb": weekly_traffic_total_tb(traffic_values),
            "prb_busy_hour_weekly_average": prb_busy_hour_weekly_average(prb_values),
            "code_drop_average": code_drop_average(code_drop_values),
            "radio_network_availability": round(average(availability_values), 4),
            "peak_throughput_mbps": round(peak_throughput, 4),
            "active_subscribers": round(active_subscribers, 4),
            "addressable_subscribers": round(addressable_subscribers, 4),
            "penetration_rate": penetration_rate(active_subscribers, addressable_subscribers),
            "average_gb_per_active_user": average_gb_per_active_user(weekly_gb, active_subscribers),
            "average_throughput_per_active_user": average_throughput_per_active_user(peak_throughput, active_subscribers),
            "excluded": site_name.lower() in excluded,
        })

    return sorted(metrics, key=lambda record: str(record["site_name"]))


def build_executive_kpis(site_metrics: Sequence[Mapping[str, object]]) -> dict[str, object]:
    included = [site for site in site_metrics if not site.get("excluded")]
    total_traffic_gb = sum(float(site.get("weekly_traffic_gb") or 0.0) for site in included)
    total_active = sum(float(site.get("active_subscribers") or 0.0) for site in included)
    total_addressable = sum(float(site.get("addressable_subscribers") or 0.0) for site in included)
    peak_throughput = maximum(float(site.get("peak_throughput_mbps") or 0.0) for site in included)

    return {
        "total_network_traffic_gb": round(total_traffic_gb, 4),
        "total_network_traffic_tb": gb_to_tb(total_traffic_gb),
        "radio_network_availability": round(average(float(site.get("radio_network_availability") or 0.0) for site in included), 4),
        "prb_utilisation": round(average(float(site.get("prb_busy_hour_weekly_average") or 0.0) for site in included), 4),
        "code_drop_rate": round(average(float(site.get("code_drop_average") or 0.0) for site in included), 4),
        "peak_throughput_mbps": round(peak_throughput, 4),
        "total_subscribers": round(total_addressable, 4),
        "active_subscribers": round(total_active, 4),
        "penetration_rate": penetration_rate(total_active, total_addressable),
        "average_gb_per_active_user": average_gb_per_active_user(total_traffic_gb, total_active),
        "average_throughput_per_active_user": average_throughput_per_active_user(peak_throughput, total_active),
    }


def build_report_sections() -> list[dict[str, str]]:
    return [
        {
            "name": "Executive Network KPI Report",
            "worksheet": "Executive KPI",
            "description": "Network-level weekly traffic, availability, PRB, code-drop, throughput, subscriber, penetration, and usage KPIs.",
            "status": "generated",
        },
        {
            "name": "GCO Report Section",
            "worksheet": "GCO Report",
            "description": "Executive KPI values arranged for GCO/GCOO copy-forward reporting.",
            "status": "generated",
        },
        {
            "name": "GCU Report Section",
            "worksheet": "GCU Report",
            "description": "Executive KPI values arranged for GCU/GCUO copy-forward reporting.",
            "status": "generated",
        },
        {
            "name": "General Reporting Section",
            "worksheet": "General Report",
            "description": "Reusable general weekly KPI report section.",
            "status": "generated",
        },
        {
            "name": "Site Performance Report",
            "worksheet": "Site Performance",
            "description": "Per-site weekly traffic, PRB, code-drop, availability, subscribers, and usage metrics.",
            "status": "generated",
        },
        {
            "name": "Traffic Rankings",
            "worksheet": "Top 20 Traffic / Bottom 20 Traffic",
            "description": "Top and bottom site rankings by weekly traffic with excluded sites removed.",
            "status": "generated",
        },
        {
            "name": "PRB Rankings",
            "worksheet": "Top 20 PRB / Bottom 20 PRB",
            "description": "Top and bottom site rankings by busy-hour PRB weekly average.",
            "status": "generated",
        },
        {
            "name": "Code Drop Rankings",
            "worksheet": "Top 20 Code Drop / Bottom 20 Code Drop",
            "description": "Top and bottom site rankings by weekly code/drop-rate average.",
            "status": "generated",
        },
        {
            "name": "Exceptions",
            "worksheet": "Exceptions",
            "description": "Non-commercialised/new/explicitly excluded sites and source rows with missing site names.",
            "status": "generated",
        },
        {
            "name": "Audit Trail",
            "worksheet": "Audit",
            "description": "Raw input reference, output path, user context, generated timestamp, and source column mapping.",
            "status": "generated",
        },
    ]


def _write_rows(sheet, rows: Sequence[Mapping[str, object]]) -> None:
    if not rows:
        sheet.append(["No data"])
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])


def _write_key_values(sheet, rows: Sequence[tuple[str, object]]) -> None:
    sheet.append(["Metric", "Value"])
    for label, value in rows:
        sheet.append([label, value])


def _autosize(sheet) -> None:
    for column in sheet.columns:
        values = [str(cell.value) for cell in column if cell.value is not None]
        width = min(max([len(value) for value in values] + [10]) + 2, 48)
        sheet.column_dimensions[column[0].column_letter].width = width


def write_report_workbook(
    output_path: Path,
    *,
    site_metrics: Sequence[Mapping[str, object]],
    top_sites: Sequence[Mapping[str, object]],
    bottom_sites: Sequence[Mapping[str, object]],
    top_prb_sites: Sequence[Mapping[str, object]],
    bottom_prb_sites: Sequence[Mapping[str, object]],
    top_code_drop_sites: Sequence[Mapping[str, object]],
    bottom_code_drop_sites: Sequence[Mapping[str, object]],
    executive_kpis: Mapping[str, object],
    sections: Sequence[Mapping[str, object]],
    exceptions: Sequence[Mapping[str, object]],
    audit_rows: Sequence[tuple[str, object]],
) -> None:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    _write_key_values(summary, [
        ("Total sites", len(site_metrics)),
        ("Included sites", sum(1 for site in site_metrics if not site.get("excluded"))),
        ("Excluded sites", sum(1 for site in site_metrics if site.get("excluded"))),
        ("Generated sections", len(sections)),
        ("Total traffic TB", executive_kpis.get("total_network_traffic_tb", 0.0)),
        ("Active subscribers", executive_kpis.get("active_subscribers", 0.0)),
        ("Penetration rate %", executive_kpis.get("penetration_rate", 0.0)),
    ])

    sections_sheet = workbook.create_sheet("Report Sections")
    _write_rows(sections_sheet, sections)

    executive_sheet = workbook.create_sheet("Executive KPI")
    _write_key_values(executive_sheet, [
        ("Total network traffic GB", executive_kpis.get("total_network_traffic_gb", 0.0)),
        ("Total network traffic TB", executive_kpis.get("total_network_traffic_tb", 0.0)),
        ("Radio network availability %", executive_kpis.get("radio_network_availability", 0.0)),
        ("PRB utilisation %", executive_kpis.get("prb_utilisation", 0.0)),
        ("Code drop rate %", executive_kpis.get("code_drop_rate", 0.0)),
        ("Peak throughput Mbps", executive_kpis.get("peak_throughput_mbps", 0.0)),
        ("Total subscribers", executive_kpis.get("total_subscribers", 0.0)),
        ("Active subscribers", executive_kpis.get("active_subscribers", 0.0)),
        ("Penetration rate %", executive_kpis.get("penetration_rate", 0.0)),
        ("Average GB per active user", executive_kpis.get("average_gb_per_active_user", 0.0)),
        ("Average throughput per active user Mbps", executive_kpis.get("average_throughput_per_active_user", 0.0)),
    ])

    for title in ["GCO Report", "GCU Report", "General Report"]:
        sheet = workbook.create_sheet(title)
        sheet.append(["Report Section", title])
        sheet.append([])
        _write_key_values(sheet, [
            ("Traffic TB", executive_kpis.get("total_network_traffic_tb", 0.0)),
            ("Availability %", executive_kpis.get("radio_network_availability", 0.0)),
            ("PRB utilisation %", executive_kpis.get("prb_utilisation", 0.0)),
            ("Code drop rate %", executive_kpis.get("code_drop_rate", 0.0)),
            ("Peak throughput Mbps", executive_kpis.get("peak_throughput_mbps", 0.0)),
            ("Total subscribers", executive_kpis.get("total_subscribers", 0.0)),
            ("Active subscribers", executive_kpis.get("active_subscribers", 0.0)),
            ("Penetration rate %", executive_kpis.get("penetration_rate", 0.0)),
            ("Average GB per active user", executive_kpis.get("average_gb_per_active_user", 0.0)),
            ("Average throughput per active user Mbps", executive_kpis.get("average_throughput_per_active_user", 0.0)),
        ])

    metrics_sheet = workbook.create_sheet("Site Performance")
    _write_rows(metrics_sheet, site_metrics)

    top_sheet = workbook.create_sheet("Top 20 Traffic")
    _write_rows(top_sheet, top_sites)

    bottom_sheet = workbook.create_sheet("Bottom 20 Traffic")
    _write_rows(bottom_sheet, bottom_sites)

    top_prb_sheet = workbook.create_sheet("Top 20 PRB")
    _write_rows(top_prb_sheet, top_prb_sites)

    bottom_prb_sheet = workbook.create_sheet("Bottom 20 PRB")
    _write_rows(bottom_prb_sheet, bottom_prb_sites)

    top_code_sheet = workbook.create_sheet("Top 20 Code Drop")
    _write_rows(top_code_sheet, top_code_drop_sites)

    bottom_code_sheet = workbook.create_sheet("Bottom 20 Code Drop")
    _write_rows(bottom_code_sheet, bottom_code_drop_sites)

    exceptions_sheet = workbook.create_sheet("Exceptions")
    _write_rows(exceptions_sheet, exceptions)

    audit_sheet = workbook.create_sheet("Audit")
    _write_key_values(audit_sheet, audit_rows)

    for sheet in workbook.worksheets:
        _autosize(sheet)

    workbook.save(output_path)


def run_report_from_file(
    source_file: Path,
    *,
    original_filename: str,
    exclusions: Iterable[str] | None = None,
    user_context: str = "local",
) -> dict[str, object]:
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    run_dir = REPORT_ROOT / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    input_path = run_dir / f"input{source_file.suffix.lower()}"
    output_path = run_dir / "netgenix_report.xlsx"
    pdf_path = run_dir / "netgenix_report.pdf"
    audit_path = run_dir / "audit.json"
    shutil.move(str(source_file), input_path)

    rows = load_tabular_file(input_path)
    for row in rows:
        row["_source_file"] = original_filename
    site_metrics = compute_site_metrics(rows, exclusions=exclusions)
    included = [site for site in site_metrics if not site.get("excluded")]
    top_sites = rank_sites(included, "weekly_traffic_gb", top_n=20, descending=True)
    bottom_sites = rank_sites(included, "weekly_traffic_gb", top_n=20, descending=False)
    top_prb_sites = rank_sites(included, "prb_busy_hour_weekly_average", top_n=20, descending=True)
    bottom_prb_sites = rank_sites(included, "prb_busy_hour_weekly_average", top_n=20, descending=False)
    top_code_drop_sites = rank_sites(included, "code_drop_average", top_n=20, descending=True)
    bottom_code_drop_sites = rank_sites(included, "code_drop_average", top_n=20, descending=False)
    executive_kpis = build_executive_kpis(site_metrics)
    sections = build_report_sections()
    exceptions = [
        {
            "site_name": site.get("site_name"),
            "reason": "Excluded from rankings",
        }
        for site in site_metrics
        if site.get("excluded")
    ]
    missing_site_rows = sum(
        1
        for row in rows
        if not str(_first(row, CONCEPT_ALIASES["site_name"], "")).strip()
    )
    if missing_site_rows:
        exceptions.append({
            "site_name": "",
            "reason": f"{missing_site_rows} source rows did not contain a recognised site name",
        })

    source_columns = sorted({str(column) for row in rows for column in row.keys()})
    audit_rows = [
        ("Run ID", run_id),
        ("Created at", datetime.now(timezone.utc).isoformat()),
        ("User context", user_context),
        ("Original filename", original_filename),
        ("Raw input file", str(input_path)),
        ("Generated output", str(output_path)),
        ("Rows imported", len(rows)),
        ("Sites computed", len(site_metrics)),
        ("Exclusions", ", ".join(exclusions or [])),
        ("Source columns", ", ".join(source_columns)),
    ]

    write_report_workbook(
        output_path,
        site_metrics=site_metrics,
        top_sites=top_sites,
        bottom_sites=bottom_sites,
        top_prb_sites=top_prb_sites,
        bottom_prb_sites=bottom_prb_sites,
        top_code_drop_sites=top_code_drop_sites,
        bottom_code_drop_sites=bottom_code_drop_sites,
        executive_kpis=executive_kpis,
        sections=sections,
        exceptions=exceptions,
        audit_rows=audit_rows,
    )
    write_pdf_report(
        pdf_path,
        run_id=run_id,
        site_metrics=site_metrics,
        top_sites=top_sites,
        bottom_sites=bottom_sites,
        top_prb_sites=top_prb_sites,
        top_code_drop_sites=top_code_drop_sites,
        executive_kpis=executive_kpis,
        original_filename=original_filename,
    )

    audit = {
        "run_id": run_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_context": user_context,
        "original_filename": original_filename,
        "raw_input_file": str(input_path),
        "generated_output": str(output_path),
        "generated_pdf": str(pdf_path),
        "site_count": len(site_metrics),
        "exclusions": list(exclusions or []),
        "sections": sections,
        "executive_kpis": executive_kpis,
        "exceptions": exceptions,
        "source_columns": source_columns,
        "computed_metrics": site_metrics,
    }
    audit_path.write_text(json.dumps(audit, indent=2), encoding="utf-8")

    return {
        "run_id": run_id,
        "input_file": str(input_path),
        "output_file": str(output_path),
        "pdf_file": str(pdf_path),
        "audit_file": str(audit_path),
        "site_count": len(site_metrics),
        "sections": sections,
        "executive_kpis": executive_kpis,
        "top_traffic_sites": top_sites,
        "bottom_traffic_sites": bottom_sites,
        "top_prb_sites": top_prb_sites,
        "bottom_prb_sites": bottom_prb_sites,
        "top_code_drop_sites": top_code_drop_sites,
        "bottom_code_drop_sites": bottom_code_drop_sites,
    }


def run_report_from_files(
    source_files: Sequence[tuple[Path, str]],
    *,
    exclusions: Iterable[str] | None = None,
    user_context: str = "local",
) -> dict[str, object]:
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    run_dir = REPORT_ROOT / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    input_dir = run_dir / "inputs"
    input_dir.mkdir(exist_ok=True)

    rows: list[dict[str, object]] = []
    input_files: list[str] = []
    original_filenames: list[str] = []

    for index, (source_file, original_filename) in enumerate(source_files, start=1):
        safe_name = "".join(ch if ch.isalnum() or ch in {".", "-", "_"} else "_" for ch in original_filename)
        input_path = input_dir / f"{index:02d}_{safe_name or ('input' + source_file.suffix.lower())}"
        shutil.move(str(source_file), input_path)
        input_files.append(str(input_path))
        original_filenames.append(original_filename)

        file_rows = load_tabular_file(input_path)
        for row in file_rows:
            row["_source_file"] = original_filename
        rows.extend(file_rows)

    output_path = run_dir / "netgenix_report.xlsx"
    pdf_path = run_dir / "netgenix_report.pdf"
    audit_path = run_dir / "audit.json"
    result = _generate_report_artifacts(
        run_id=run_id,
        rows=rows,
        output_path=output_path,
        pdf_path=pdf_path,
        audit_path=audit_path,
        input_file_label=", ".join(input_files),
        original_filename=", ".join(original_filenames),
        exclusions=exclusions,
        user_context=user_context,
    )
    result["input_file"] = ", ".join(input_files)
    return result


def run_report_from_rows(
    rows: Sequence[Mapping[str, object]],
    *,
    source_label: str,
    exclusions: Iterable[str] | None = None,
    user_context: str = "automation",
    evaluation_only: bool = True,
) -> dict[str, object]:
    """Generate report artifacts from normalized database rows."""
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    run_dir = REPORT_ROOT / run_id
    run_dir.mkdir(parents=True, exist_ok=False)
    result = _generate_report_artifacts(
        run_id=run_id,
        rows=rows,
        output_path=run_dir / "netgenix_report.xlsx",
        pdf_path=run_dir / "netgenix_report.pdf",
        audit_path=run_dir / "audit.json",
        input_file_label=source_label,
        original_filename=source_label,
        exclusions=exclusions,
        user_context=user_context,
        evaluation_only=evaluation_only,
    )
    return result


def _generate_report_artifacts(
    *,
    run_id: str,
    rows: Sequence[Mapping[str, object]],
    output_path: Path,
    pdf_path: Path,
    audit_path: Path,
    input_file_label: str,
    original_filename: str,
    exclusions: Iterable[str] | None,
    user_context: str,
    evaluation_only: bool = False,
) -> dict[str, object]:
    rows = _deduplicate_rows(rows)
    site_metrics = compute_site_metrics(rows, exclusions=exclusions)
    included = [site for site in site_metrics if not site.get("excluded")]
    top_sites = rank_sites(included, "weekly_traffic_gb", top_n=20, descending=True)
    bottom_sites = rank_sites(included, "weekly_traffic_gb", top_n=20, descending=False)
    top_prb_sites = rank_sites(included, "prb_busy_hour_weekly_average", top_n=20, descending=True)
    bottom_prb_sites = rank_sites(included, "prb_busy_hour_weekly_average", top_n=20, descending=False)
    top_code_drop_sites = rank_sites(included, "code_drop_average", top_n=20, descending=True)
    bottom_code_drop_sites = rank_sites(included, "code_drop_average", top_n=20, descending=False)
    executive_kpis = build_executive_kpis(site_metrics)
    if evaluation_only:
        for key in (
            "peak_throughput_mbps",
            "total_subscribers",
            "active_subscribers",
            "penetration_rate",
            "average_gb_per_active_user",
            "average_throughput_per_active_user",
        ):
            executive_kpis[key] = "N/A"
    sections = build_report_sections()
    exceptions = [
        {
            "site_name": site.get("site_name"),
            "reason": "Excluded from rankings",
        }
        for site in site_metrics
        if site.get("excluded")
    ]
    missing_site_rows = sum(
        1
        for row in rows
        if not str(_first(row, CONCEPT_ALIASES["site_name"], "")).strip()
    )
    if missing_site_rows:
        exceptions.append({
            "site_name": "",
            "reason": f"{missing_site_rows} source rows did not contain a recognised site name",
        })

    source_columns = sorted({str(column) for row in rows for column in row.keys()})
    audit_rows = [
        ("Run ID", run_id),
        ("Created at", datetime.now(timezone.utc).isoformat()),
        ("User context", user_context),
        ("Original filename", original_filename),
        ("Raw input file", input_file_label),
        ("Generated output", str(output_path)),
        ("Rows imported", len(rows)),
        ("Sites computed", len(site_metrics)),
        ("Exclusions", ", ".join(exclusions or [])),
        ("Source columns", ", ".join(source_columns)),
    ]

    write_report_workbook(
        output_path,
        site_metrics=site_metrics,
        top_sites=top_sites,
        bottom_sites=bottom_sites,
        top_prb_sites=top_prb_sites,
        bottom_prb_sites=bottom_prb_sites,
        top_code_drop_sites=top_code_drop_sites,
        bottom_code_drop_sites=bottom_code_drop_sites,
        executive_kpis=executive_kpis,
        sections=sections,
        exceptions=exceptions,
        audit_rows=audit_rows,
    )
    write_pdf_report(
        pdf_path,
        run_id=run_id,
        site_metrics=site_metrics,
        top_sites=top_sites,
        bottom_sites=bottom_sites,
        top_prb_sites=top_prb_sites,
        top_code_drop_sites=top_code_drop_sites,
        executive_kpis=executive_kpis,
        original_filename=original_filename,
    )

    audit = {
        "run_id": run_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_context": user_context,
        "original_filename": original_filename,
        "raw_input_file": input_file_label,
        "generated_output": str(output_path),
        "generated_pdf": str(pdf_path),
        "site_count": len(site_metrics),
        "exclusions": list(exclusions or []),
        "sections": sections,
        "executive_kpis": executive_kpis,
        "exceptions": exceptions,
        "source_columns": source_columns,
        "computed_metrics": site_metrics,
    }
    audit_path.write_text(json.dumps(audit, indent=2), encoding="utf-8")

    return {
        "run_id": run_id,
        "input_file": input_file_label,
        "output_file": str(output_path),
        "pdf_file": str(pdf_path),
        "audit_file": str(audit_path),
        "site_count": len(site_metrics),
        "sections": sections,
        "executive_kpis": executive_kpis,
        "top_traffic_sites": top_sites,
        "bottom_traffic_sites": bottom_sites,
        "top_prb_sites": top_prb_sites,
        "bottom_prb_sites": bottom_prb_sites,
        "top_code_drop_sites": top_code_drop_sites,
        "bottom_code_drop_sites": bottom_code_drop_sites,
    }


def get_report_output_path(run_id: str) -> Path:
    return REPORT_ROOT / run_id / "netgenix_report.xlsx"


def get_report_pdf_path(run_id: str) -> Path:
    return REPORT_ROOT / run_id / "netgenix_report.pdf"


def list_report_runs() -> list[dict[str, object]]:
    if not REPORT_ROOT.exists():
        return []

    runs = []
    for audit_path in sorted(REPORT_ROOT.glob("*/audit.json"), reverse=True):
        try:
            audit = json.loads(audit_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue

        run_id = str(audit.get("run_id") or audit_path.parent.name)
        pdf_path = Path(str(audit.get("generated_pdf") or get_report_pdf_path(run_id)))
        pdf_exists = pdf_path.exists()
        runs.append({
            "run_id": run_id,
            "created_at": audit.get("created_at"),
            "original_filename": audit.get("original_filename"),
            "site_count": int(audit.get("site_count") or 0),
            "sections_count": len(audit.get("sections") or []),
            "output_file": str(audit.get("generated_output") or get_report_output_path(run_id)),
            "download_url": f"/api/reports/runs/{run_id}/download",
            "pdf_file": str(pdf_path) if pdf_exists else None,
            "pdf_download_url": f"/api/reports/runs/{run_id}/download/pdf" if pdf_exists else None,
            "audit_file": str(audit_path),
        })

    return runs
